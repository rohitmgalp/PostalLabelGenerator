import streamlit as st
import openpyxl
import os
import sys
import json
import time
import io
import base64
import re
import requests
import pandas as pd
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont

# --- SYSTEM DIRECTORY SETUP ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "web_postal_data.json")

DISPATCH_ARTICLES = [
    "Speed Post Parcel", 
    "Indiapost Parcel Retail", 
    "Business Parcel", 
    "Speed Post Parcel COD", 
    "Business Parcel COD"
]

BARCODE_POOL_KEYS = [
    "Speed Post Parcel (Regular & COD Shared Pool)",
    "Indiapost Parcel Retail",
    "Business Parcel",
    "Business Parcel COD"
]

# --- OMNIPRESENT DATABASE HEALER ---
def load_data():
    default_db = {"users": {}, "messages": []}
    if not os.path.exists(DATA_FILE): 
        return default_db
    try:
        with open(DATA_FILE, "r") as f:
            content = f.read().strip()
            if not content: return default_db
            data = json.loads(content)
            if not isinstance(data, dict): return default_db
            
            # This completely eliminates KeyErrors. It forces the structure to exist.
            if "users" not in data: data["users"] = {}
            if "messages" not in data: data["messages"] = []
            
            for uid, udata in data["users"].items():
                if "used_barcodes" not in udata: udata["used_barcodes"] = []
                if "generated_labels" not in udata: udata["generated_labels"] = []
                if "addresses" not in udata: udata["addresses"] = []
                if "barcodes" not in udata: udata["barcodes"] = {}
                for pk in BARCODE_POOL_KEYS:
                    if pk not in udata["barcodes"]:
                        udata["barcodes"][pk] = {"prefix": "", "current": 0, "end": 0, "suffix": ""}
            
            return data
    except Exception:
        return default_db

def save_data(data):
    with open(DATA_FILE, "w") as f: json.dump(data, f)

def get_pool_key(article_type):
    if article_type in ["Speed Post Parcel", "Speed Post Parcel COD"]:
        return "Speed Post Parcel (Regular & COD Shared Pool)"
    return article_type

# --- UPU S10 COMPLIANT CHECK DIGIT ENGINE ---
def calculate_upu_s10_check_digit(serial_8_digits):
    serial_str = f"{int(serial_8_digits):08d}"
    digits = [int(d) for d in serial_str]
    weights = [8, 6, 4, 2, 3, 5, 9, 7]
    total_sum = sum(d * w for d, w in zip(digits, weights))
    remainder = total_sum % 11
    c = 11 - remainder
    if c == 10: return "0"
    elif c == 11: return "5"
    else: return str(c)

# --- TEXT WRAPPING ENGINE ---
def wrap_text_to_pixels(text, draw, font, max_width):
    text_str = str(text)
    lines = []
    for p in text_str.split('\n'):
        words = p.split(' ')
        current_line = ""
        for word in words:
            test_line = current_line + " " + word if current_line else word
            bbox = draw.textbbox((0, 0), test_line, font=font)
            w = bbox[2] - bbox[0]
            if w <= max_width: 
                current_line = test_line
            else:
                if current_line: lines.append(current_line)
                current_line = word
        if current_line: lines.append(current_line)
    return '\n'.join(lines)

# --- INTELLIGENT DATA EXTRACTION ENGINE ---
def extract_pincode_and_mobile(text):
    pincode, mobile = "", ""
    if not text: return pincode, mobile
    chunks = re.findall(r'\+?\d+', text)
    for chunk in chunks:
        digits_only = chunk.replace('+', '')
        if len(digits_only) == 6: pincode = digits_only
        elif len(digits_only) == 10: mobile = digits_only
        elif len(digits_only) in [11, 12, 13]: mobile = digits_only[-10:]
    return pincode, mobile

# --- BULLETPROOF OFFLINE CSV LOADER ---
@st.cache_data(show_spinner=False)
def load_pincode_database_records():
    csv_path = os.path.join(BASE_DIR, "all_india_pincode_directory_2025.csv")
    if not os.path.exists(csv_path): return {}
    try:
        df = pd.read_csv(csv_path, dtype=str, on_bad_lines='skip')
        df.columns = [str(c).lower().strip() for c in df.columns]
        if 'pincode' not in df.columns or 'district' not in df.columns or 'statename' not in df.columns: return {}
        df['pincode'] = df['pincode'].fillna("").astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
        df = df[df['pincode'] != ""] 
        df_unique = df.drop_duplicates(subset=['pincode'], keep='first').fillna("")
        return df_unique.set_index('pincode').to_dict(orient='index')
    except Exception: return {}

def safe_numeric(val):
    if val is None: return None
    s = str(val).strip()
    if not s: return None
    try:
        if '.' in s: return float(s)
        return int(s)
    except: return s

def split_address_to_lines(address_text):
    lines = [line.strip() for line in str(address_text).split('\n') if line.strip()]
    name = lines[0] if len(lines) > 0 else "CUSTOMER"
    l1 = lines[1] if len(lines) > 1 else ""
    l2 = lines[2] if len(lines) > 2 else ""
    l3 = ", ".join(lines[3:]) if len(lines) > 3 else ""
    if not l1 and address_text:
        flat = str(address_text).replace('\n', ', ')
        name = flat[:40]
        l1 = flat[40:90]
        l2 = flat[90:140]
        l3 = flat[140:190]
    return name, l1, l2, l3

# --- LABEL GENERATOR ---
def draw_single_label(entry, width_in, height_in):
    DPI = 300
    W_px = int(width_in * DPI)
    H_px = int(height_in * DPI)
    if not entry or not isinstance(entry, dict):
        return Image.new('RGB', (W_px, H_px), color='white')
        
    m_px = int(W_px * 0.05)
    tracking_code = str(entry.get('tracking', '0000000000000'))
    article_type = str(entry.get('article', 'ARTICLE'))
    
    Code128 = barcode.get_barcode_class('code128')
    bc_buffer = io.BytesIO()
    my_barcode = Code128(tracking_code, writer=ImageWriter())
    my_barcode.write(bc_buffer, options={"write_text": False, "background": "white", "quiet_zone": 1.0})
    bc_buffer.seek(0)
    bc_img = Image.open(bc_buffer)
    
    lbl_canvas = Image.new('RGB', (W_px, H_px), color='white')
    draw = ImageDraw.Draw(lbl_canvas)
    scale = min(W_px, H_px)
    
    size_l = max(36, int(scale * 0.058))
    size_m = max(28, int(scale * 0.044))
    size_b = max(32, int(scale * 0.048))
    size_bc_text = max(30, int(scale * 0.046))
    
    try:
        f_l = ImageFont.load_default(size=size_l)
        f_m = ImageFont.load_default(size=size_m)
        f_b = ImageFont.load_default(size=size_b)
        f_bc = ImageFont.load_default(size=size_bc_text)
    except:
        f_l = f_m = f_b = f_bc = ImageFont.load_default()
        
    top_strings = [f"ARTICLE: {article_type}"]
    if entry.get('cust_id'): top_strings.append(f"CUST ID: {entry.get('cust_id')}")
    if entry.get('cod'): top_strings.append(f"COD CHARGES: Rs. {entry.get('cod')}")
    
    if W_px >= H_px:
        bc_h_scaled = int(H_px * 0.13)
        bc_w_scaled = int(W_px * 0.65)
        bc_y_pos = H_px - bc_h_scaled - m_px - int(H_px * 0.05)
        col_width = (W_px - (3 * m_px)) // 2
        y_left = m_px
        for line in top_strings:
            draw.text((m_px, y_left), line, fill="black", font=f_b)
            b_box = draw.textbbox((0,0), line, font=f_b)
            y_left += (b_box[3] - b_box[1]) + int(H_px * 0.015)
        y_left += int(H_px * 0.02)
        draw.text((m_px, y_left), "FROM:", fill="black", font=f_b)
        y_left += int(size_b * 1.2)
        w_from = wrap_text_to_pixels(entry.get('from', ''), draw, f_m, col_width)
        draw.multiline_text((m_px, y_left), w_from, fill="black", font=f_m, spacing=8)
        
        x_right = m_px + col_width + m_px
        y_right = m_px
        draw.text((x_right, y_right), "TO:", fill="black", font=f_b)
        y_right += int(size_b * 1.2)
        w_to = wrap_text_to_pixels(entry.get('to', ''), draw, f_l, col_width)
        draw.multiline_text((x_right, y_right), w_to, fill="black", font=f_l, spacing=8)
        
        bc_resized = bc_img.resize((bc_w_scaled, bc_h_scaled))
        lbl_canvas.paste(bc_resized, ((W_px - bc_w_scaled) // 2, bc_y_pos))
    else: 
        use_w = W_px - (2 * m_px)
        bc_h_scaled = int(H_px * 0.11)
        bc_w_scaled = int(W_px * 0.75)
        bc_y_pos = H_px - bc_h_scaled - m_px - int(H_px * 0.05)
        y_curr = m_px
        for line in top_strings:
            draw.text((m_px, y_curr), line, fill="black", font=f_b)
            b_box = draw.textbbox((0,0), line, font=f_b)
            y_curr += (b_box[3] - b_box[1]) + int(H_px * 0.012)
        y_curr += int(H_px * 0.02)
        draw.text((m_px, y_curr), "FROM:", fill="black", font=f_b)
        y_curr += int(size_b * 1.2)
        w_from = wrap_text_to_pixels(entry.get('from', ''), draw, f_m, use_w)
        draw.multiline_text((m_px, y_curr), w_from, fill="black", font=f_m, spacing=8)
        b_box = draw.multiline_textbbox((m_px, y_curr), w_from, font=f_m, spacing=8)
        y_cursor_from = b_box[3] + int(H_px * 0.04)
        draw.text((m_px, y_cursor_from), "TO:", fill="black", font=f_b)
        y_cursor_from += int(size_b * 1.2)
        w_to = wrap_text_to_pixels(entry.get('to', ''), draw, f_l, use_w)
        draw.multiline_text((m_px, y_cursor_from), w_to, fill="black", font=f_l, spacing=8)
        bc_resized = bc_img.resize((bc_w_scaled, bc_h_scaled))
        lbl_canvas.paste(bc_resized, ((W_px - bc_w_scaled) // 2, bc_y_pos))
        
    bbox_bc = draw.textbbox((0, 0), tracking_code, font=f_bc)
    text_bc_w = bbox_bc[2] - bbox_bc[0]
    draw.text(((W_px - text_bc_w) // 2, bc_y_pos + bc_h_scaled + int(H_px * 0.012)), tracking_code, fill="black", font=f_bc)
    return lbl_canvas

def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# --- STYLING & SETTINGS ---
st.set_page_config(page_title="India Post Enterprise Workspace", page_icon="📮", layout="wide")
bg_file_path = os.path.join(BASE_DIR, "background.png")

whatsapp_html = """
    <a href="https://wa.me/918075386388" target="_blank" class="whatsapp-float">
        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" fill="currentColor" viewBox="0 0 16 16" style="margin-right: 8px; display: inline-block; vertical-align: middle;">
            <path d="M13.601 2.326A7.85 7.85 0 0 0 7.994 0C3.627 0 .068 3.558.064 7.926c0 1.399.366 2.76 1.057 3.965L0 16l4.204-1.102a7.9 7.9 0 0 0 3.79.965h.004c4.368 0 7.926-3.558 7.93-7.93A7.9 7.9 0 0 0 13.6 2.326zM7.994 14.521a6.6 6.6 0 0 1-3.356-.92l-.24-.144-2.494.654.666-2.433-.156-.251a6.56 6.56 0 0 1-1.007-3.505c1.11-4.285 4.564-6.574 8.718-6.574a6.58 6.58 0 0 1 4.66 2.72 6.58 6.58 0 0 1 1.936 4.663c-.004 4.2-3.563 7.759-7.923 7.759M11.57 9.447c-.19-.094-1.127-.556-1.301-.62-.174-.064-.3-.094-.426.094-.126.188-.488.62-.6.749-.113.128-.226.144-.417.05-.19-.095-.807-.296-1.536-.855-.567-.457-.951-1.022-1.062-1.116-.112-.094-.012-.145.083-.242.085-.087.174-.188.26-.283.085-.087.174-.188.26-.283.087-.095.116-.16.174-.319.059-.158.03-.3-.015-.394-.045-.094-.426-1.026-.583-1.409-.153-.367-.307-.317-.418-.317-.109-.004-.234-.004-.36-.004a.69.69 0 0 0-.5.234c-.174.188-.665.65-0.665 1.583s.678 1.834.773 1.96c.095.127 1.332 2.035 3.226 2.856.45.195.8.311 1.075.398.452.144.863.124 1.189.062.363-.069 1.127-.461 1.284-.906.158-.444.158-.825.11-1.013-.048-.19-.174-.3-.365-.394"/>
        </svg>Contact Us
    </a>
"""

if os.path.exists(bg_file_path):
    encoded_bg = get_base64_image(bg_file_path)
    st.markdown(f"""
        <style>
            .stApp {{ background-image: url("data:image/png;base64,{encoded_bg}"); background-size: cover; background-position: center; background-attachment: fixed; }}
            div[data-testid="stHeader"] {{ background: transparent !important; }}
            .stTextInput input, .stTextArea textarea, .stSelectbox div {{ border-color: #cbd5e1 !important; background-color: #ffffff !important; color: #0f172a !important; }}
            div.stButton > button[type="primary"] {{ background-color: #9c0000 !important; color: white !important; border: none !important; font-weight: bold !important; padding: 10px 24px !important; border-radius: 8px !important; }}
            div.stButton > button[type="primary"]:hover {{ background-color: #bd0000 !important; }}
            .whatsapp-float {{
                position: fixed; bottom: 24px; right: 24px; background-color: #25d366; color: white !important;
                padding: 12px 22px; border-radius: 30px; text-decoration: none !important; font-family: 'Segoe UI', sans-serif;
                font-weight: 700; font-size: 14px; box-shadow: 0 4px 15px rgba(0,0,0,0.22); z-index: 99999;
                display: inline-flex; align-items: center; transition: transform 0.2s ease, background-color 0.2s ease;
            }}
            .whatsapp-float:hover {{ background-color: #1ebd58; transform: scale(1.05); }}
        </style>
        {whatsapp_html}
    """, unsafe_allow_html=True)
else:
    st.markdown(f"""
        <style>
            .stApp {{ background-color: #fdfbf7; }}
            div.stButton > button[type="primary"] {{ background-color: #9c0000 !important; color: white !important; }}
            .whatsapp-float {{
                position: fixed; bottom: 24px; right: 24px; background-color: #25d366; color: white !important;
                padding: 12px 22px; border-radius: 30px; text-decoration: none !important; font-family: sans-serif;
                font-weight: 700; font-size: 14px; box-shadow: 0 4px 15px rgba(0,0,0,0.22); z-index: 99999;
                display: inline-flex; align-items: center; transition: transform 0.2s ease;
            }}
            .whatsapp-float:hover {{ transform: scale(1.05); }}
        </style>
        {whatsapp_html}
    """, unsafe_allow_html=True)

# --- STRICT SESSION STATE INIT ---
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
if 'username' not in st.session_state: st.session_state.username = ""
if 'web_queue' not in st.session_state: st.session_state.web_queue = []
if 's_id' not in st.session_state: st.session_state.s_id = 0
if 'r_id' not in st.session_state: st.session_state.r_id = 0
if 'stage_err' not in st.session_state: st.session_state.stage_err = ""
if 'stage_succ' not in st.session_state: st.session_state.stage_succ = ""
if 'load_profile_dd' not in st.session_state: st.session_state.load_profile_dd = "-- Select Profile --"

# --- SAFE CALLBACKS ---
def s_addr_changed():
    sid = st.session_state.s_id
    addr = st.session_state.get(f"s_addr_{sid}", "")
    _, mob = extract_pincode_and_mobile(addr)
    if mob: st.session_state[f"s_mob_{sid}"] = mob

def r_addr_changed():
    rid = st.session_state.r_id
    addr = st.session_state.get(f"r_addr_{rid}", "")
    pin, mob = extract_pincode_and_mobile(addr)
    if pin: st.session_state[f"r_pin_{rid}"] = pin
    if mob: st.session_state[f"r_mob_{rid}"] = mob

def load_profile():
    choice = st.session_state.load_profile_dd
    if choice != "-- Select Profile --":
        st.session_state.s_id += 1
        new_sid = st.session_state.s_id
        st.session_state[f"s_addr_{new_sid}"] = choice
        _, mob = extract_pincode_and_mobile(choice)
        if mob: st.session_state[f"s_mob_{new_sid}"] = mob

def execute_stage(tracking, article, pool_key, current_serial):
    sid = st.session_state.s_id
    rid = st.session_state.r_id
    
    from_val = st.session_state.get(f"s_addr_{sid}", "").strip()
    to_val = st.session_state.get(f"r_addr_{rid}", "").strip()
    
    if not from_val or not to_val or not tracking:
        st.session_state.stage_err = "From Address, To Address, and Tracking ID are mandatory."
        return
        
    current_u = st.session_state.username
    db = load_data()
    
    # Failsafe if user was deleted mid-session
    if current_u not in db["users"]:
        st.session_state.stage_err = "Critical Error: User profile missing. Please log out."
        return

    st.session_state.web_queue.append({
        "tracking": tracking, "from": from_val, "to": to_val, "article": article,
        "cod": st.session_state.get(f"cod_{rid}", "").strip(),
        "cust_id": st.session_state.get("cust_shared", "").strip(),
        "weight": st.session_state.get(f"w_{rid}", "").strip(),
        "length": st.session_state.get(f"l_{rid}", "").strip(),
        "breadth": st.session_state.get(f"b_{rid}", "").strip(),
        "height": st.session_state.get(f"h_{rid}", "").strip(),
        "s_mob": st.session_state.get(f"s_mob_{sid}", "").strip(),
        "r_mob": st.session_state.get(f"r_mob_{rid}", "").strip(),
        "pincode": st.session_state.get(f"r_pin_{rid}", "").strip()
    })
    
    db["users"][current_u]["used_barcodes"].append(tracking)
    db["users"][current_u]["barcodes"][pool_key]["current"] = current_serial + 1
    save_data(db)
    
    st.session_state.r_id += 1 
    st.session_state.stage_succ = "Staged successfully! Ready for the next recipient."

pincode_lookup_db = load_pincode_database_records()

# --- AUTHENTICATION ---
if not st.session_state.authenticated:
    st.markdown("""
        <div style="text-align: left; margin-top: 15px; margin-bottom: 25px; font-family: 'Segoe UI', system-ui, sans-serif;">
            <h1 style="color: #9c0000; font-size: 3.4rem; font-weight: 800; margin: 0; line-height: 1.1; letter-spacing: -0.5px;">India Post</h1>
            <h2 style="color: #334155; font-size: 1.9rem; font-weight: 600; margin-top: 4px; margin-bottom: 8px; opacity: 0.95;">Enterprise Web Portal</h2>
        </div>
        <div style="height: 1px; background: rgba(156, 0, 0, 0.15); margin-bottom: 30px;"></div>
    """, unsafe_allow_html=True)

    auth_cols = st.columns([0.4, 0.6])
    with auth_cols[0]:
        with st.container(border=True):
            auth_mode = st.radio("Access Control Node", ["Login to Existing Profile", "Register New Corporate Profile"])
            if auth_mode == "Login to Existing Profile":
                st.markdown("<h4 style='color:#9c0000; margin-top:10px;'>🔐 Sign In</h4>", unsafe_allow_html=True)
                user_id = st.text_input("User ID").strip()
                password = st.text_input("Password", type="password").strip()
                if st.button("Verify & Enter Portal", type="primary", use_container_width=True):
                    if not user_id or not password:
                        st.error("Please enter both User ID and Password fields.")
                    else:
                        data = load_data()
                        if user_id in data.get("users", {}) and data["users"][user_id]["password"] == password:
                            if data["users"][user_id].get("status", "active") == "locked":
                                st.error("❌ Access Exception: Locked.")
                            else:
                                st.session_state.authenticated = True
                                st.session_state.username = user_id
                                st.rerun()
                        else:
                            st.error("Invalid credentials.")
            else:
                st.markdown("<h4 style='color:#9c0000; margin-top:10px;'>📝 Register</h4>", unsafe_allow_html=True)
                reg_name = st.text_input("Full Name / Company Name").strip()
                reg_email = st.text_input("Email ID").strip()
                reg_mobile = st.text_input("Mobile Number").strip()
                user_id = st.text_input("Create User ID").strip()
                password = st.text_input("Create Password", type="password").strip()
                if st.button("Register", type="primary", use_container_width=True):
                    if not reg_name or not reg_email or not reg_mobile or not user_id or not password:
                        st.error("All fields mandatory.")
                    else:
                        data = load_data()
                        if user_id in data.get("users", {}): st.error("Occupied ID.")
                        else:
                            data["users"][user_id] = {
                                "name": reg_name, "email": reg_email, "mobile": reg_mobile, "password": password,
                                "status": "active", "addresses": [], "used_barcodes": [], "generated_labels": [],
                                "barcodes": {k: {"prefix": "", "current": 0, "end": 0, "suffix": ""} for k in BARCODE_POOL_KEYS}
                            }
                            save_data(data)
                            st.success("Success! Please log in.")
    st.stop()

# --- SIDEBAR DIAGNOSTICS ---
if pincode_lookup_db:
    st.sidebar.success(f"✅ Secure CSV Database Loaded: **{len(pincode_lookup_db)}** routes active.")
else:
    st.sidebar.error("⚠️ CSV Pincode Database NOT loaded! Ensure 'all_india_pincode_directory_2025.csv' is uploaded.")

# --- LOAD DATA & MESSAGES ---
current_user = st.session_state.username
db = load_data()

if current_user not in db["users"]:
    st.session_state.authenticated = False
    st.session_state.username = ""
    st.rerun()

user_profile = db["users"][current_user]

# --- DASHBOARD HEADER ---
col_logout_wrap = st.columns([0.80, 0.20])
with col_logout_wrap[0]:
    st.markdown(f"<h4 style='margin:0; color:#1e293b;'>📋 Active Client: {user_profile.get('name', current_user)} | Node ID: `{current_user}`</h4>", unsafe_allow_html=True)
with col_logout_wrap[1]:
    if st.button("Core Log Out", use_container_width=True):
        st.session_state.clear()
        st.rerun()

# --- NOTIFICATIONS ---
if current_user.lower() != "admin":
    my_unread_msgs = [m for m in db.get("messages", []) if m.get("to") == current_user and m.get("status") == "unread"]
    if my_unread_msgs:
        st.markdown("<h4 style='color: #b45309; margin-bottom: 0px;'>🔔 Important Notifications</h4>", unsafe_allow_html=True)
        for msg in my_unread_msgs:
            with st.container(border=True):
                st.warning(f"**Admin Message:** {msg['text']}")
                reply_input = st.text_input("Send a Reply to Admin:", key=f"reply_in_{msg['id']}")
                col1, col2 = st.columns([0.2, 0.8])
                with col1:
                    if st.button("Send Reply", key=f"btn_rep_{msg['id']}"):
                        if reply_input.strip():
                            msg["reply"] = reply_input.strip()
                            msg["status"] = "replied"
                            save_data(db)
                            st.rerun()
                with col2:
                    if st.button("Mark as Read (No Reply)", key=f"btn_rd_{msg['id']}"):
                        msg["status"] = "read"
                        save_data(db)
                        st.rerun()

# --- TABS ---
tabs_list = ["📋 Dispatch Manager", "⚙️ Settings & Barcode Ranges", "📇 Generated Labels"]
if current_user.lower() == "admin": 
    tabs_list.append("👥 Admin Directory")
    tabs_list.append("✉️ Message Center")
tabs = st.tabs(tabs_list)

# --- TAB 1: DISPATCH MANAGER ---
with tabs[0]:
    col_inputs, col_preview = st.columns([0.48, 0.52])
    
    with col_inputs:
        with st.container(border=True):
            st.markdown("<h4 style='color:#9c0000; margin-top:0;'>📦 Shipment Properties</h4>", unsafe_allow_html=True)
            col_w_in, col_h_in = st.columns(2)
            with col_w_in: width_in = st.number_input("Label Width (Inches)", value=6.0, step=0.5)
            with col_h_in: height_in = st.number_input("Label Height (Inches)", value=4.0, step=0.5)
            
            s_id = st.session_state.s_id
            r_id = st.session_state.r_id
            
            saved_addresses = user_profile.get("addresses", [])
            st.selectbox("Quick-Load Saved 'From' Address", ["-- Select Profile --"] + saved_addresses, key="load_profile_dd", on_change=load_profile)
            
            st.text_area("Sender 'From' Address Details", key=f"s_addr_{s_id}", on_change=s_addr_changed)
            
            col_addr_actions = st.columns(2)
            with col_addr_actions[0]:
                if st.button("💾 Remember Address", use_container_width=True):
                    val = st.session_state.get(f"s_addr_{s_id}", "").strip()
                    if val and val not in user_profile["addresses"]:
                        db["users"][current_user]["addresses"].append(val)
                        save_data(db)
                        st.success("Address profile recorded.")
                        st.rerun()
            with col_addr_actions[1]:
                if st.button("🗑️ Delete Address", use_container_width=True):
                    val = st.session_state.load_profile_dd
                    if val != "-- Select Profile --" and val in user_profile["addresses"]:
                        db["users"][current_user]["addresses"].remove(val)
                        save_data(db)
                        st.session_state.load_profile_dd = "-- Select Profile --"
                        st.warning("Address profile removed.")
                        st.rerun()
            
            st.text_area("Recipient 'To' Address Details", key=f"r_addr_{r_id}", on_change=r_addr_changed)
            
            article_type = st.selectbox("Postal Article Class", DISPATCH_ARTICLES)
            
            if "COD" in article_type: 
                st.text_input("Collect on Delivery (COD) Amount (₹)", key=f"cod_{r_id}")
            st.text_input("India Post Customer Business ID", key="cust_shared")
            
            st.write("**Volumetric Specifications (Optional)**")
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            with col_m1: st.text_input("Weight (g)", key=f"w_{r_id}")
            with col_m2: st.text_input("Len (cm)", key=f"l_{r_id}")
            with col_m3: st.text_input("Wid (cm)", key=f"b_{r_id}")
            with col_m4: st.text_input("Hgt (cm)", key=f"h_{r_id}")
                
            col_mob1, col_mob2 = st.columns(2)
            with col_mob1: st.text_input("Sender Mobile (Optional)", key=f"s_mob_{s_id}")
            with col_mob2: st.text_input("Receiver Mobile (Optional)", key=f"r_mob_{r_id}")
                
            col_pin1, col_pin2 = st.columns(2)
            with col_pin1: st.text_input("Extracted Pincode (Optional)", key=f"r_pin_{r_id}")
            with col_pin2: st.write("")

            shared_pool_key = get_pool_key(article_type)
            b_current = user_profile["barcodes"][shared_pool_key]
            used_set = set(user_profile.get("used_barcodes", []))
            queue_set = {item.get("tracking") for item in st.session_state.web_queue if isinstance(item, dict)}

            if b_current["current"] == 0 or b_current["current"] > b_current["end"]:
                st.error(f"❌ Shared series empty! Configure ranges for: {shared_pool_key}")
                auto_tracking = None
                current_serial_8 = 0
            else:
                current_serial_8 = int(b_current["current"])
                auto_tracking = None
                while current_serial_8 <= int(b_current["end"]):
                    check_digit = calculate_upu_s10_check_digit(current_serial_8)
                    test_tracking = f"{b_current['prefix']}{current_serial_8:08d}{check_digit}{b_current['suffix']}"
                    if test_tracking not in used_set and test_tracking not in queue_set:
                        auto_tracking = test_tracking
                        break
                    current_serial_8 += 1
                if auto_tracking: st.info(f"Next S10 Tracking ID: **{auto_tracking}**")
                else: st.error("❌ Barcode Pool Depleted! Update configuration strings.")

            if st.session_state.stage_err:
                st.error(st.session_state.stage_err)
                st.session_state.stage_err = ""
            if st.session_state.stage_succ:
                st.success(st.session_state.stage_succ)
                st.session_state.stage_succ = ""

            st.button("➕ Stage to Batch Allocation Queue", type="primary", on_click=execute_stage, args=(auto_tracking, article_type, shared_pool_key, current_serial_8))

    with col_preview:
        with st.container(border=True):
            st.markdown("<h4 style='color:#9c0000; margin-top:0;'>⏳ Staged Processing Queue</h4>", unsafe_allow_html=True)
            if st.session_state.web_queue:
                display_df = pd.DataFrame([item for item in st.session_state.web_queue if isinstance(item, dict)])[["tracking", "article", "weight", "cust_id"]]
                st.dataframe(display_df, use_container_width=True)
                if st.button("🗑️ Wipe Entire Active Session Queue"):
                    st.session_state.web_queue = []
                    st.rerun()
                    
                st.write("---")
                st.subheader("Compile Outputs")
                
                if st.button("⚙️ Compile Label PDFs & Sync Template Matrix", type="primary"):
                    template_filename = os.path.join(BASE_DIR, "New Format bulk.xlsx")
                    if not os.path.exists(template_filename): template_filename = os.path.join(BASE_DIR, "Template_Master.xlsx")
                        
                    if not os.path.exists(template_filename):
                        st.error("CRITICAL: Master template tracking sheet asset missing from directory.")
                    else:
                        with st.spinner("Compiling manifest locally against database..."):
                            pdf_pages = []
                            wb = openpyxl.load_workbook(template_filename)
                            ws = wb.active
                            next_row = ws.max_row + 1
                            
                            for idx, entry in enumerate(st.session_state.web_queue):
                                if not isinstance(entry, dict): continue
                                    
                                lbl_canvas = draw_single_label(entry, width_in, height_in)
                                if lbl_canvas: pdf_pages.append(lbl_canvas)
                                
                                # OFFLINE CSV SECURE LOOKUP
                                r_pin_clean = str(entry.get('pincode', '')).strip().split('.')[0]
                                if not r_pin_clean:
                                    r_pin_clean, _ = extract_pincode_and_mobile(entry.get('to', ''))
                                    r_pin_clean = str(r_pin_clean).strip().split('.')[0]
                                r_pin_details = pincode_lookup_db.get(r_pin_clean)
                                if not isinstance(r_pin_details, dict): r_pin_details = {}
                                r_name, r_l1, _, _ = split_address_to_lines(entry.get('to', ''))
                                
                                s_pin, _ = extract_pincode_and_mobile(entry.get('from', ''))
                                s_pin_clean = str(s_pin).strip().split('.')[0]
                                s_pin_details = pincode_lookup_db.get(s_pin_clean)
                                if not isinstance(s_pin_details, dict): s_pin_details = {}
                                _, s_l1, s_l2, _ = split_address_to_lines(entry.get('from', ''))
                                
                                # EXCEL INJECTIONS
                                ws.cell(row=next_row, column=1, value=idx + 1)
                                ws.cell(row=next_row, column=2, value=entry.get('tracking', ''))
                                ws.cell(row=next_row, column=3, value=safe_numeric(entry.get('weight', '')))
                                ws.cell(row=next_row, column=4, value="FALSE")
                                ws.cell(row=next_row, column=5, value="FALSE")
                                ws.cell(row=next_row, column=6, value=str(r_pin_details.get('district', '')).upper())
                                ws.cell(row=next_row, column=7, value=r_pin_clean)
                                ws.cell(row=next_row, column=8, value=r_name)
                                ws.cell(row=next_row, column=9, value=r_l1)
                                ws.cell(row=next_row, column=10, value=str(r_pin_details.get('district', '')).upper())
                                ws.cell(row=next_row, column=11, value=str(r_pin_details.get('statename', '')).upper())
                                ws.cell(row=next_row, column=12, value="FALSE")
                                ws.cell(row=next_row, column=13, value=entry.get('s_mob', ''))
                                ws.cell(row=next_row, column=14, value=entry.get('r_mob', ''))
                                
                                if "COD" in entry.get('article', ''):
                                    ws.cell(row=next_row, column=17, value="COD")
                                    ws.cell(row=next_row, column=18, value=safe_numeric(entry.get('cod', '')))
                                    
                                ws.cell(row=next_row, column=21, value="NROL")
                                ws.cell(row=next_row, column=22, value=safe_numeric(entry.get('length', '')))
                                ws.cell(row=next_row, column=23, value=safe_numeric(entry.get('breadth', '')))
                                ws.cell(row=next_row, column=24, value=safe_numeric(entry.get('height', '')))
                                ws.cell(row=next_row, column=25, value="FALSE")
                                ws.cell(row=next_row, column=29, value=user_profile.get('name', current_user))
                                ws.cell(row=next_row, column=31, value=str(s_pin_details.get('district', '')).upper())
                                ws.cell(row=next_row, column=32, value=str(s_pin_details.get('statename', '')).upper())
                                ws.cell(row=next_row, column=33, value=s_pin_clean)
                                ws.cell(row=next_row, column=39, value=str(r_pin_details.get('statename', '')).upper())
                                ws.cell(row=next_row, column=44, value="FALSE")
                                ws.cell(row=next_row, column=45, value="RMGK REF")
                                ws.cell(row=next_row, column=46, value=s_l1)
                                ws.cell(row=next_row, column=47, value=s_l2)
                                ws.cell(row=next_row, column=48, value=str(s_pin_details.get('statename', '')).upper())
                                
                                next_row += 1
                                user_profile["generated_labels"].append(entry)
                                
                            db["users"][current_user] = user_profile
                            save_data(db)
                            
                            if pdf_pages:
                                pdf_buffer = io.BytesIO()
                                pdf_pages[0].save(pdf_buffer, "PDF", save_all=True, append_images=pdf_pages[1:], resolution=300.0)
                                st.session_state.pdf_ready = pdf_buffer.getvalue()
                            
                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            st.session_state.excel_ready = excel_buffer.getvalue()
                            st.session_state.web_queue = [] 
                            st.success("Compilation complete! Web download links active.")
                            st.rerun()
            else:
                st.info("The dispatch pipeline queue is currently clean.")

            if 'pdf_ready' in st.session_state or 'excel_ready' in st.session_state:
                st.write("---")
                st.markdown("<h5 style='color:#9c0000; margin-top:0;'>📥 Generated Files Cache</h5>", unsafe_allow_html=True)
                batch_timestamp = int(time.time())
                
                if 'pdf_ready' in st.session_state:
                    st.download_button("📥 Download Consolidated Label PDF Bundle", data=st.session_state.pdf_ready, file_name=f"Compiled_Labels_{batch_timestamp}.pdf", mime="application/pdf", use_container_width=True)
                if 'excel_ready' in st.session_state:
                    st.download_button("📥 Download Template Upload Ready Manifest", data=st.session_state.excel_ready, file_name=f"Bulk_Upload_{batch_timestamp}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                
                if st.button("🧹 Clear Download Cache History", use_container_width=True):
                    if 'pdf_ready' in st.session_state: del st.session_state.pdf_ready
                    if 'excel_ready' in st.session_state: del st.session_state.excel_ready
                    st.rerun()

# --- TAB 2: RANGE SETTINGS ---
with tabs[1]:
    with st.container(border=True):
        st.markdown("<h4 style='color:#9c0000; margin-top:0;'>⚙️ UPU S10 Barcode Range Setup</h4>", unsafe_allow_html=True)
        set_article = st.selectbox("Choose Target Allocation Track Key", BARCODE_POOL_KEYS)
        b_data = user_profile["barcodes"][set_article]
        
        col_p, col_st, col_en, col_su = st.columns(4)
        with col_p: new_prefix = st.text_input("Prefix (2 Alpha)", value=b_data["prefix"], max_chars=2).upper()
        with col_st: new_start = st.number_input("Start Serial (8 Digits)", value=int(b_data["current"]), step=1)
        with col_en: new_end = st.number_input("End Serial (8 Digits)", value=int(b_data["end"]), step=1)
        with col_su: new_suffix = st.text_input("Suffix (2 Alpha)", value=b_data["suffix"], max_chars=2).upper()
        
        if st.button("Save System Range Allocation", type="primary"):
            db["users"][current_user]["barcodes"][set_article] = { "prefix": new_prefix, "current": new_start, "end": new_end, "suffix": new_suffix }
            save_data(db)
            st.success(f"Configured range assignment arrays for {set_article}!")
            st.rerun()
        st.metric("Available Unused Serials Remaining", value=f"{max(0, b_data['end'] - b_data['current'] + 1) if b_data['end'] > 0 else 0} Units")

# --- TAB 3: PERMANENT ARCHIVE ---
with tabs[2]:
    with st.container(border=True):
        st.markdown("<h4 style='color:#9c0000; margin-top:0;'>📇 Permanent Generated Labels Registry</h4>", unsafe_allow_html=True)
        archive = user_profile.get("generated_labels", [])
        if not archive: st.info("No labels generated yet.")
        else:
            for idx, item in enumerate(reversed(archive)):
                real_idx = len(archive) - 1 - idx
                row_cols = st.columns([0.5, 0.25, 0.25])
                with row_cols[0]:
                    st.write(f"**Barcode:** `{item['tracking']}` | **Type:** {item['article']}")
                    st.caption(f"**Recipient:** {item['to'].splitlines()[0][:35]}...")
                with row_cols[1]:
                    lbl_canvas = draw_single_label(item, 6.0, 4.0)
                    if lbl_canvas:
                        reprint_buf = io.BytesIO()
                        lbl_canvas.save(reprint_buf, "PDF", resolution=300.0)
                        st.download_button("🖨️ Reprint PDF", data=reprint_buf.getvalue(), file_name=f"Reprint_{item['tracking']}.pdf", mime="application/pdf", key=f"rep_{item['tracking']}_{real_idx}")
                with row_cols[2]:
                    if st.button(f"🗑️ Delete Record", key=f"del_init_{real_idx}"): st.session_state[f"confirm_prompt_{real_idx}"] = True
                
                if st.session_state.get(f"confirm_prompt_{real_idx}", False):
                    st.markdown(f"❓ **Do you want to reuse barcode `{item['tracking']}`?**")
                    choice_cols = st.columns([0.3, 0.3, 0.4])
                    with choice_cols[0]:
                        if st.button("Yes (Return to Range)", key=f"re_yes_{real_idx}"):
                            if item['tracking'] in user_profile["used_barcodes"]: user_profile["used_barcodes"].remove(item['tracking'])
                            user_profile["generated_labels"].pop(real_idx)
                            db["users"][current_user] = user_profile
                            save_data(db)
                            del st.session_state[f"confirm_prompt_{real_idx}"]
                            st.rerun()
                    with choice_cols[1]:
                        if st.button("No (Burn Barcode)", key=f"re_no_{real_idx}"):
                            user_profile["generated_labels"].pop(real_idx)
                            db["users"][current_user] = user_profile
                            save_data(db)
                            del st.session_state[f"confirm_prompt_{real_idx}"]
                            st.rerun()
                    with choice_cols[2]:
                        if st.button("Cancel Operations", key=f"re_can_{real_idx}"):
                            del st.session_state[f"confirm_prompt_{real_idx}"]
                            st.rerun()
                st.markdown("<hr style='margin:10px 0; border-color:rgba(156,0,0,0.15);'>", unsafe_allow_html=True)

# --- TABS 4 & 5: ADMIN CONTROLS ---
if current_user.lower() == "admin":
    with tabs[3]:
        with st.container(border=True):
            st.markdown("<h4 style='color:#9c0000; margin-top:0;'>👥 Corporate Client Infrastructure Directory</h4>", unsafe_allow_html=True)
            admin_db = load_data()
            for uid, info in list(admin_db.get("users", {}).items()):
                if uid.lower() == "admin": continue
                u_status = info.get("status", "active")
                adm_row = st.columns([0.34, 0.22, 0.22, 0.22])
                with adm_row[0]:
                    st.markdown(f"**User ID:** `{uid}` | **Name:** {info.get('name','N/A')}")
                    st.caption(f"📧 **Email:** {info.get('email', 'N/A')} | 📱 **Mobile:** {info.get('mobile', 'N/A')}")
                    st.caption(f"🟢 State: `{u_status.upper()}`")
                with adm_row[1]:
                    if st.button("🔑 Reset Password", key=f"adm_pwd_{uid}", use_container_width=True):
                        admin_db["users"][uid]["password"] = "123456"
                        save_data(admin_db)
                        st.rerun()
                with adm_row[2]:
                    if u_status == "active":
                        if st.button("🔒 Lock User", key=f"adm_lock_{uid}", use_container_width=True):
                            admin_db["users"][uid]["status"] = "locked"
                            save_data(admin_db)
                            st.rerun()
                    else:
                        if st.button("🔓 Unlock User", key=f"adm_unl_{uid}", use_container_width=True):
                            admin_db["users"][uid]["status"] = "active"
                            save_data(admin_db)
                            st.rerun()
                with adm_row[3]:
                    if st.button("🚨 Delete User", key=f"adm_del_{uid}", use_container_width=True):
                        del admin_db["users"][uid]
                        save_data(admin_db)
                        st.rerun()
                st.markdown("<hr style='margin:12px 0; border-color:rgba(0,0,0,0.06);'>", unsafe_allow_html=True)
                
    with tabs[4]:
        with st.container(border=True):
            st.markdown("<h4 style='color:#9c0000; margin-top:0;'>✉️ Broadcast Message Center</h4>", unsafe_allow_html=True)
            all_users = [u for u in db.get("users", {}).keys() if u.lower() != "admin"]
            target = st.selectbox("Select Message Recipient", ["All Users"] + all_users)
            msg_text = st.text_area("Message Content")
            if st.button("Send Message", type="primary"):
                if msg_text.strip():
                    targets = all_users if target == "All Users" else [target]
                    for t in targets:
                        db["messages"].append({
                            "id": f"msg_{int(time.time()*1000)}_{t}",
                            "to": t,
                            "text": msg_text,
                            "reply": "",
                            "status": "unread",
                            "timestamp": time.time()
                        })
                    save_data(db)
                    st.success("Message successfully dispatched!")
                    st.rerun()
                else:
                    st.error("Message cannot be blank.")
                    
            st.markdown("---")
            col_msg_head1, col_msg_head2 = st.columns([0.75, 0.25])
            with col_msg_head1: st.markdown("### 📥 User Replies & History")
            with col_msg_head2:
                if st.button("🗑️ Clear All History", type="primary"):
                    db["messages"] = []
                    save_data(db)
                    st.rerun()
                    
            messages_list = db.get("messages", [])
            if not messages_list:
                st.info("No messages sent yet.")
            else:
                for i, m in enumerate(reversed(messages_list)):
                    real_idx = len(messages_list) - 1 - i
                    with st.container(border=True):
                        col_m1, col_m2 = st.columns([0.85, 0.15])
                        with col_m1:
                            st.markdown(f"**To User Node:** `{m.get('to')}` | **Status:** `{m.get('status', 'unread').upper()}`")
                            st.info(f"**Admin Sent:** {m.get('text')}")
                            if m.get("reply"):
                                st.success(f"**Reply Received:** {m.get('reply')}")
                        with col_m2:
                            if st.button("❌ Delete", key=f"del_msg_ind_{real_idx}"):
                                db["messages"].pop(real_idx)
                                save_data(db)
                                st.rerun()
